#!/usr/bin/env python3
"""
生成铁三训练计划 Excel
Zikun - Sprint Triathlon 2026-05-30
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

OUTPUT = "/root/.openclaw/workspace/training/triathlon_plan.xlsx"

# ── 颜色定义 ──────────────────────────────────────────────
C_SWIM   = "BDD7EE"   # 蓝
C_BIKE   = "E2EFDA"   # 绿
C_RUN    = "FCE4D6"   # 橙
C_BRICK  = "F4B942"   # 金色（骑跑砖训）
C_REST   = "F2F2F2"   # 灰（休息）
C_RACE   = "FF0000"   # 红（比赛日）
C_HEADER = "2F5496"   # 深蓝表头
C_WEEK   = "D6E4F0"   # 周行背景

# ── 10周每日训练计划 ──────────────────────────────────────
# 格式: (运动类型, 时间窗口, 训练内容, 心率区间, 备注)
# 类型: swim/bike/run/brick/rest/race

PLAN = [
  # W1: 3/23-3/29  直接开干（已有3月训练基础）
  {
    "week": 1, "label": "W1 直接开干", "dates": ("2026-03-23", "2026-03-29"),
    "focus": "已有3月基础，W1直接上强度。节奏跑+阈值段落，游泳结构化，周末Brick骑行提心率至Z3",
    "days": [
      ("run",   "早7:30-9:00",  "节奏跑 6km（热身1km+节奏4km @6:20-6:40/km+放松1km）", "Z2-Z3",  "心率目标Z3(130-148)，乳酸阈值下方"),
      ("swim",  "早7:30-9:00",  "4×200m（组间30s）+ 200m计时（记录成绩）",              "Z2-Z3",  "200m目标<6min，感受配速"),
      ("run",   "早7:30-9:00",  "阈值间歇：3×1km @Z4（148-167bpm），组间慢跑90s",        "Z4",     "配速目标5:45-6:10/km，比乳酸阈值配速稍慢"),
      ("rest",  "—",            "休息日",                                                "—",      "充分恢复，拉伸/冥想"),
      ("run",   "午12:00-14:00","轻松跑 6km（Z2，主动恢复）",                            "Z2",     "比昨天间歇后的恢复跑，保持轻松"),
      ("brick", "周末全天",      "🧱砖训：骑30km（HR 140-155，Z3-Z4）→ 立即跑5km（目标6:40-7:00）","骑Z3-Z4/跑Z3","骑行HR要比3/22的129高！感受比赛强度"),
      ("swim",  "早7:30-9:00",  "🎯750m基准计时（连续，全程记录时间）",                  "Z3",     "建立基准，后续各周对比进步"),
    ]
  },
  # W2: 3/30-4/5  强度巩固+游泳超距
  {
    "week": 2, "label": "W2 强度巩固", "dates": ("2026-03-30", "2026-04-05"),
    "focus": "跑步加入长距节奏跑，游泳超距800m，骑行Brick提速",
    "days": [
      ("run",   "早7:30-9:00",  "长跑 8km（全程Z2，稳定7:00-7:30配速）",                "Z2",     "有氧基础，别追配速，保持心率Z2"),
      ("swim",  "早7:30-9:00",  "2×400m（Z3）+ 100m放松",                               "Z3+Z2",  "400m目标<12min"),
      ("run",   "早7:30-9:00",  "节奏跑 6km（热身1km+节奏4km @6:20-6:40+放松1km）",     "Z2-Z3",  ""),
      ("rest",  "—",            "休息日",                                                "—",      ""),
      ("run",   "午12:00-14:00","阈值间歇：4×1km @Z4，组间慢跑90s",                      "Z4",     "比W1多一组，配速目标5:45-6:10"),
      ("brick", "周末全天",      "🧱砖训：骑35km（HR 143-155，Z3-Z4）→ 立即跑5km（目标6:30-6:45）","骑Z3-Z4/跑Z3","骑行距离提升5km，跑步配速提升"),
      ("swim",  "早7:30-9:00",  "超距：800m连续游",                                      "Z2-Z3",  "不停顿，对比750m基准感受余量"),
    ]
  },
  # W3: 4/6-4/12  骑行专项强化
  {
    "week": 3, "label": "W3 骑行强化", "dates": ("2026-04-06", "2026-04-12"),
    "focus": "骑行强度最强周（含高强度区间），跑步维持阈值，游泳1000m超距",
    "days": [
      ("run",   "早7:30-9:00",  "节奏跑 6km（热身1km+节奏4km @6:10-6:30+放松1km）",     "Z3",     "比W2配速再提5-10s"),
      ("swim",  "早7:30-9:00",  "超距：1000m连续游",                                     "Z2-Z3",  "目标<31min，建立信心"),
      ("run",   "早7:30-9:00",  "阈值间歇：5×1km @Z4，组间慢跑90s",                      "Z4",     "本周最高强度跑，配速5:40-6:05"),
      ("rest",  "—",            "休息日",                                                "—",      ""),
      ("run",   "午12:00-14:00","轻松跑 6km（Z2，恢复）",                                "Z2",     ""),
      ("bike",  "周末全天",      "骑行 40km（含20km @ Z3-Z4，HR 143-160）",               "Z2-Z4",  "前10km热身，中20km强度，末10km Z2收尾"),
      ("swim",  "早7:30-9:00",  "4×200m（Z3）+ 200m放松",                                "Z3",     ""),
    ]
  },
  # W4: 4/13-4/19  整合测试
  {
    "week": 4, "label": "W4 整合测试", "dates": ("2026-04-13", "2026-04-19"),
    "focus": "750m游泳计时测试（对比W1基准），完整Brick升级为骑跑全距，跑步配速推进",
    "days": [
      ("run",   "早7:30-9:00",  "节奏跑 7km（热身1km+节奏5km @6:10-6:30+放松1km）",     "Z2-Z3",  ""),
      ("swim",  "早7:30-9:00",  "🎯750m计时测试（对比W1成绩）",                           "Z3-Z4",  "记录成绩，应比W1快"),
      ("run",   "早7:30-9:00",  "轻松跑 6km（Z2，游泳测试后恢复）",                       "Z2",     ""),
      ("rest",  "—",            "休息日",                                    "—",              ""),
      ("run",   "午12:00-14:00","轻松跑 5km + 4×200m配速跑",                 "Z2+Z4",          "200m配速跑约1:10-1:15"),
      ("brick", "周末全天",      "🧱砖训：骑30km（HR 143-155，Z3-Z4）→ 跑5km（目标6:30-6:45）","骑Z3-Z4/跑Z3","骑行HR要比上周高，感受比赛强度"),
      ("swim",  "早7:30-9:00",  "4×200m（Z3）+ 100m放松",                   "Z3",             ""),
    ]
  },
  # W5: 4/20-4/26  训练高峰
  {
    "week": 5, "label": "W5 训练高峰", "dates": ("2026-04-20", "2026-04-26"),
    "focus": "全周训练量最高峰，骑行40km，游泳超距1000m，跑步节奏跑+阈值双刺激",
    "days": [
      ("run",   "早7:30-9:00",  "长跑 8km（Z2，稳定7:00-7:30配速）",         "Z2",             "有氧基础长跑，不追配速"),
      ("swim",  "早7:30-9:00",  "超距：1000m连续游",                          "Z2-Z3",          "目标<31min，建立750m余量信心"),
      ("run",   "早7:30-9:00",  "阈值间歇：5×1km @Z4，组间90s",               "Z4",             "高峰周最强跑步，配速5:35-6:00"),
      ("rest",  "—",            "休息日",                                     "—",              "高峰周中段充分恢复"),
      ("run",   "午12:00-14:00","节奏跑 5km（热身1km+节奏3km @Z3+放松1km）",   "Z2-Z3",          ""),
      ("bike",  "周末全天",      "骑行 40km（含25km @ Z3，HR 130-148）",        "Z2-Z3",          "全程最长骑行，后半段维持节奏"),
      ("swim",  "早7:30-9:00",  "🎯750m计时（对比W4成绩）",                    "Z3-Z4",          "记录成绩，应有进步"),
    ]
  },
  # W6: 4/27-5/3  首次完整三项模拟
  {
    "week": 6, "label": "W6 首次全程模拟", "dates": ("2026-04-27", "2026-05-03"),
    "focus": "首次完整三项模拟，练习换项（T1/T2），跑步继续保持节奏跑/阈值强度",
    "days": [
      ("run",   "早7:30-9:00",  "节奏跑 6km（热身1km+节奏4km @6:00-6:20+放松1km）","Z3",     "配速比W5再提"),
      ("swim",  "早7:30-9:00",  "游泳换项演练：750m游完→出水换装计时",       "Z3",             "T1目标<2min，反复练出水动作"),
      ("run",   "早7:30-9:00",  "阈值间歇：4×1km @Z4，组间90s",              "Z4",             ""),
      ("rest",  "—",            "休息日",                                    "—",              ""),
      ("run",   "午12:00-14:00","轻松跑 4km（Z2，为周末模拟保存体能）",        "Z2",             ""),
      ("brick", "周末全天",      "🏁首次完整模拟：游750m→T1→骑20km→T2→跑5km", "比赛强度Z3",    "全程计时！记录各段，感受节奏"),
      ("bike",  "周末全天",      "轻松骑 20km（主动恢复）",                   "Z2",             "昨日全程模拟后恢复"),
    ]
  },
  # W7: 5/4-5/10  强化完整模拟
  {
    "week": 7, "label": "W7 强化完整模拟", "dates": ("2026-05-04", "2026-05-10"),
    "focus": "完整比赛距离第二次模拟，跑步配速目标6:30，骑行提速至26-27km/h",
    "days": [
      ("run",   "早7:30-9:00",  "节奏跑 7km（热身1km+节奏5km @5:55-6:15+放松1km）","Z3",     ""),
      ("swim",  "早7:30-9:00",  "🎯750m计时+换项演练（对比W6）",              "Z3-Z4",          "记录成绩，应有进步"),
      ("run",   "早7:30-9:00",  "阈值间歇：5×1km @Z4，组间90s",              "Z4",             "配速5:35-6:00/km"),
      ("rest",  "—",            "休息日",                                    "—",              ""),
      ("run",   "午12:00-14:00","轻松跑 4km（Z2，为周末模拟保存体能）",        "Z2",             ""),
      ("brick", "周末全天",      "🏁完整模拟②：游750m→T1→骑20km（提速26+km/h）→T2→跑5km（目标6:30）","比赛强度Z3-Z4","对比W6全程时间，应有进步"),
      ("run",   "早7:30-9:00",  "恢复慢跑 3km",                              "Z1-Z2",          ""),
    ]
  },
  # W8: 5/11-5/17  减量一（-30%）
  {
    "week": 8, "label": "W8 减量一（-30%）", "dates": ("2026-05-11", "2026-05-17"),
    "focus": "开始减量，训练量-30%，保持强度感觉不消失，让身体超量恢复",
    "days": [
      ("run",   "早7:30-9:00",  "节奏跑 5km（热身1km+节奏3km @Z3+放松1km）", "Z2-Z3",          "缩短距离，保持强度"),
      ("swim",  "早7:30-9:00",  "600m（Z2，感受水感）",                       "Z2",             ""),
      ("run",   "早7:30-9:00",  "阈值间歇：3×1km @Z4，组间90s",              "Z4",             "组数减少，强度不降"),
      ("rest",  "—",            "休息日",                                    "—",              ""),
      ("run",   "午12:00-14:00","轻松跑 4km",                                 "Z2",             ""),
      ("bike",  "周末全天",      "骑行 25km（Z2-Z3，轻松强度）",               "Z2-Z3",          ""),
      ("swim",  "早7:30-9:00",  "400m轻松游",                                 "Z2",             ""),
    ]
  },
  # W9: 5/18-5/24  减量二（-50%）+ 赛前测试
  {
    "week": 9, "label": "W9 减量二（-50%）", "dates": ("2026-05-18", "2026-05-24"),
    "focus": "训练量-50%，赛前最后测试确认状态，保持肌肉激活",
    "days": [
      ("run",   "早7:30-9:00",  "轻松跑 3km + 4×200m提速跑",                 "Z2+Z3",          ""),
      ("swim",  "早7:30-9:00",  "🎯赛前750m计时测试",                         "Z3",             "对比各周成绩，确认进步"),
      ("run",   "早7:30-9:00",  "轻松跑 3km",                                "Z2",             ""),
      ("rest",  "—",            "休息日",                                    "—",              ""),
      ("run",   "午12:00-14:00","🎯赛前5km配速跑（目标<33min）",               "Z3",             ""),
      ("bike",  "周末全天",      "🎯赛前20km骑行计时（目标<46min）",            "Z3",             ""),
      ("rest",  "—",            "休息+装备最终确认",                          "—",              "检查装备，熟悉赛道"),
    ]
  },
  # W10: 5/25-5/30  赛前激活
  {
    "week": 10, "label": "W10 赛前激活", "dates": ("2026-05-25", "2026-05-30"),
    "focus": "极轻量训练保腿感，5/28-29完全休息，5/30比赛日全力以赴",
    "days": [
      ("swim",  "早7:30-9:00",  "轻松游 300m + 4×50m感受配速",               "Z1-Z2",          ""),
      ("run",   "早7:30-9:00",  "轻松跑 2km + 6×100m提速跑",                  "Z1-Z2",          "激活腿部肌肉"),
      ("bike",  "早7:30-9:00",  "轻松骑 20min + 3×1min提速",                  "Z1-Z2",          "检查装备，确认变速正常"),
      ("rest",  "—",            "完全休息",                                   "—",              "早睡，保证充足睡眠"),
      ("rest",  "—",            "完全休息，装备打包",                          "—",              "补充碳水，不尝试新食物"),
      ("race",  "比赛日",        "🏁 RACE DAY！750m游+20km骑+5km跑",          "全力",            "目标：2小时内完赛！享受每一刻！"),
      None,
    ]
  },
]

SPORT_COLORS = {
    "swim":  C_SWIM,
    "bike":  C_BIKE,
    "run":   C_RUN,
    "brick": C_BRICK,
    "rest":  C_REST,
    "race":  C_RACE,
}

SPORT_EMOJI = {
    "swim":  "🏊",
    "bike":  "🚴",
    "run":   "🏃",
    "brick": "🧱",
    "rest":  "😴",
    "race":  "🏁",
}

DAY_NAMES = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def write_cell(ws, row, col, value, bold=False, fill=None, font_color="000000",
               size=10, wrap=True, align="left", valign="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=size)
    cell.alignment = Alignment(wrap_text=wrap, horizontal=align, vertical=valign)
    cell.border = thin_border()
    if fill:
        cell.fill = make_fill(fill)
    return cell

def generate():
    wb = openpyxl.Workbook()

    # ── Sheet 1: 每日训练计划 ────────────────────────────────
    ws = wb.active
    ws.title = "每日训练计划"
    ws.freeze_panes = "B3"

    # 标题行
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value="🏊🚴🏃 Zikun 铁三备赛训练计划  |  目标：2026-05-30 Sprint Triathlon  |  目标完赛 < 2小时")
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.fill = make_fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    ws.row_dimensions[1].height = 28

    # 表头
    headers = ["周次/重点", "星期", "日期", "训练项目", "时间窗口", "训练内容", "心率区间", "备注"]
    col_widths = [22, 7, 12, 8, 14, 45, 16, 28]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        write_cell(ws, 2, i, h, bold=True, fill=C_HEADER,
                   font_color="FFFFFF", size=10, align="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 18

    row = 3
    for week_data in PLAN:
        week_start_row = row
        days = week_data["days"]

        for di, day in enumerate(days):
            if day is None:
                row += 1
                continue
            sport, window, content, hr, note = day
            day_name = DAY_NAMES[di]
            color = SPORT_COLORS.get(sport, "FFFFFF")
            emoji = SPORT_EMOJI.get(sport, "")

            # 周次列（第一天才写，后续行合并）
            write_cell(ws, row, 1, "", fill=C_WEEK)
            write_cell(ws, row, 2, day_name, fill=color, align="center")
            write_cell(ws, row, 3, "", fill=color, align="center")
            write_cell(ws, row, 4, emoji + " " + sport.upper(), fill=color, align="center", bold=(sport=="race"))
            write_cell(ws, row, 5, window, fill=color, align="center")
            write_cell(ws, row, 6, content, fill=color)
            write_cell(ws, row, 7, hr, fill=color, align="center")
            write_cell(ws, row, 8, note, fill=color)
            ws.row_dimensions[row].height = 32
            row += 1

        # 合并周次列
        ws.merge_cells(start_row=week_start_row, start_column=1,
                       end_row=row-1, end_column=1)
        mc = ws.cell(row=week_start_row, column=1)
        mc.value = f"{week_data['label']}\n{week_data['dates'][0]}～{week_data['dates'][1]}\n\n📌 {week_data['focus']}"
        mc.font = Font(bold=True, size=9, color="1F3864")
        mc.fill = make_fill(C_WEEK)
        mc.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
        mc.border = thin_border()

        # 添加日期
        start_date = date.fromisoformat(week_data["dates"][0])
        r = week_start_row
        for di, day in enumerate(days):
            if day is None:
                continue
            d = start_date + timedelta(days=di)
            ws.cell(row=r, column=3).value = d.strftime("%m/%d")
            r += 1

        # 周间分隔
        row += 0

    # ── Sheet 2: 阶段目标总览 ───────────────────────────────
    ws2 = wb.create_sheet("阶段目标总览")
    ws2.freeze_panes = "B3"

    ws2.merge_cells("A1:E1")
    c2 = ws2.cell(row=1, column=1, value="📊 各阶段训练目标总览")
    c2.font = Font(bold=True, color="FFFFFF", size=13)
    c2.fill = make_fill(C_HEADER)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = thin_border()
    ws2.row_dimensions[1].height = 28

    headers2 = ["周次", "🏊 游泳重点", "🚴 骑行重点", "🏃 跑步重点", "整周核心目标"]
    widths2 = [20, 35, 35, 35, 30]
    for i, (h, w) in enumerate(zip(headers2, widths2), 1):
        write_cell(ws2, 2, i, h, bold=True, fill=C_HEADER, font_color="FFFFFF", align="center")
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[2].height = 18

    overview = [
        ("W1 3/23-3/29", "结构化组次，划水节奏重建",       "25km有氧骑，踏频85rpm",         "轻松跑5km×3，建立节奏",       "平稳重启，负荷温和"),
        ("W2 3/30-4/5",  "4×200m，换气节奏；600m基准测试", "30km骑，引入砖训（骑后跑3km）",  "引入节奏跑，含Z3段落",        "骑跑砖训首次引入"),
        ("W3 4/6-4/12",  "超距800m，2×400m强度组",         "35km，含15km Z3区间段",          "阈值跑4×1km @Z4，长跑6km",   "专项强化，骑跑提强度"),
        ("W4 4/13-4/19", "🎯750m基准计时测试",               "砖训强度骑（HR 143-155）",       "配速推进，目标6:30-6:45",     "游泳基准测试，砖训升级"),
        ("W5 4/20-4/26", "超距1000m + 750m计时",            "最长骑行40km",                   "长跑8km + 节奏跑5km双刺激",  "训练量高峰周"),
        ("W6 4/27-5/3",  "换项演练，游后出水换装",           "首次完整三项模拟（游500+骑20+跑3）","砖训跑配速6:30",            "首次完整三项模拟"),
        ("W7 5/4-5/10",  "750m计时+换项",                   "完整距离：游750+骑20+跑5",        "骑后跑5km，稳定6:30-7:00",  "完整比赛距离演练"),
        ("W8 5/11-5/17", "600m轻松游，保感觉",               "20km轻骑（Z2）",                 "轻松跑4km×2，节奏跑4km",     "减量-30%，超量恢复"),
        ("W9 5/18-5/24", "🎯750m赛前计时",                   "🎯20km赛前计时",                  "🎯5km配速跑",                "减量-50%，赛前测试"),
        ("W10 5/25-5/30","300m保感觉（5/28停）",             "20min轻骑（5/28停）",            "2km慢跑（5/28停）",          "5/30 🏁 比赛日！"),
    ]

    row_fills = [C_SWIM, "FFFFFF"] * 10
    for i, (wk, sw, bk, rn, core) in enumerate(overview, 3):
        fill = "F7FBFF" if i % 2 == 1 else "FFFFFF"
        if i == 12:  # W10 比赛周
            fill = "FFF0F0"
        write_cell(ws2, i, 1, wk, bold=True, fill=fill, align="center")
        write_cell(ws2, i, 2, sw, fill=C_SWIM)
        write_cell(ws2, i, 3, bk, fill=C_BIKE)
        write_cell(ws2, i, 4, rn, fill=C_RUN)
        write_cell(ws2, i, 5, core, bold=(i==12), fill=fill, align="center")
        ws2.row_dimensions[i].height = 40

    # ── Sheet 3: 关键数据 ────────────────────────────────────
    ws3 = wb.create_sheet("个人数据&心率区间")

    ws3.merge_cells("A1:C1")
    c3 = ws3.cell(row=1, column=1, value="📋 Zikun 个人数据 & 心率区间参考")
    c3.font = Font(bold=True, color="FFFFFF", size=13)
    c3.fill = make_fill(C_HEADER)
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.border = thin_border()
    ws3.row_dimensions[1].height = 28

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 35

    personal = [
        ("指标", "数值", "说明"),
        ("年龄", "42岁", ""),
        ("身高/体重", "183cm / 81kg", ""),
        ("VO2 Max", "44 ml/kg/min", "良好水平（同龄均值38-42）"),
        ("实测最大心率", "185 bpm", "3/3、3/5实测"),
        ("乳酸阈值心率", "172 bpm", "对应配速5:27/km"),
        ("静息心率", "55 bpm", "运动员水平"),
        ("游泳配速", "3:00/100m", "750m预计用时约22-23min"),
        ("耐力得分", "5192", "3个月从4606增长+12.7%"),
        ("", "", ""),
        ("心率区间", "范围（bpm）", "训练用途"),
        ("Z1 恢复", "< 111", "极轻松，积极恢复"),
        ("Z2 有氧", "111–130", "80%训练量的核心区间"),
        ("Z3 节奏", "130–148", "铁三比赛配速区间"),
        ("Z4 阈值", "148–167", "乳酸阈值训练（172验证）"),
        ("Z5 最大", "167–185", "短时间冲刺"),
        ("", "", ""),
        ("比赛时间预估", "", ""),
        ("🏊 游泳750m", "~22-23 min", "@3:00/100m"),
        ("T1 换项", "~2 min", "目标<2min"),
        ("🚴 骑行20km", "~46-48 min", "@25-26km/h"),
        ("T2 换项", "~1 min", "目标<1min"),
        ("🏃 跑步5km", "~33-35 min", "@6:30-7:00/km"),
        ("🏁 总计", "~104-109 min", "目标：< 120分钟"),
    ]

    for i, (a, b, c_val) in enumerate(personal, 2):
        is_header = a in ("指标", "心率区间", "比赛时间预估")
        fill = C_HEADER if is_header else ("F7FBFF" if i % 2 == 0 else "FFFFFF")
        fc = "FFFFFF" if is_header else "000000"
        write_cell(ws3, i, 1, a, bold=is_header, fill=fill, font_color=fc)
        write_cell(ws3, i, 2, b, bold=is_header, fill=fill, font_color=fc, align="center")
        write_cell(ws3, i, 3, c_val, fill=fill)
        ws3.row_dimensions[i].height = 20

    wb.save(OUTPUT)
    print(f"✅ Excel 已生成：{OUTPUT}")

if __name__ == "__main__":
    generate()
