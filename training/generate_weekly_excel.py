#!/usr/bin/env python3
"""
每周日运行：分析本周训练数据，生成未来两周详细Excel
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/root/.openclaw/workspace/training/weekly_plan_detail.xlsx"

C_SWIM   = "BDD7EE"
C_BIKE   = "E2EFDA"
C_RUN    = "FCE4D6"
C_BRICK  = "FFE699"
C_REST   = "F2F2F2"
C_HEADER = "2F5496"
C_ACTUAL = "E2EFDA"  # 实际完成
C_WARN   = "FFC7CE"  # 偏差提醒

def tb():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(c): return PatternFill("solid", fgColor=c)

def wc(ws, row, col, val, bold=False, fc="000000", bg=None, sz=10,
       wrap=True, ha="left", va="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fc, size=sz)
    c.alignment = Alignment(wrap_text=wrap, horizontal=ha, vertical=va)
    c.border = tb()
    if bg: c.fill = fill(bg)
    return c

# ════════════════════════════════════════════════════════════
# 本周实际完成（3/16-3/22）
# ════════════════════════════════════════════════════════════
ACTUAL_WEEK = {
    "label": "W0 实际完成（3/16-3/22）",
    "summary": "本周完成6次训练，骑行量大增（含Brick），游泳3次但强度偏低（3/20 HR=101均值），跑步2次，配速偏慢",
    "highlights": [
        "✅ 3/21 骑行30.7km（86min，HR124，踏频84）— 有氧基础骑",
        "✅ 3/22 Brick：骑20km(54min,HR129) + T1(1:55) + 跑5km(35min,HR157,pace6:56) — 首次Brick，跑步HR偏高！",
        "⚠️ 3/20 游泳575m（66min，HR101）— 强度极低，可能练习/休息为主",
        "⚠️ 跑步配速 7:40/km(3/16) 和 9:10/km(3/18，跑步机低速？) — 需提升配速",
        "⚠️ 3/22 Brick跑步 HR157（Z4），配速6:56 — 骑后跑步心率明显偏高，换项疲劳大",
        "💡 骑行HR普遍偏低(124-129)，下周需要有意识提升至Z3(130-148)",
    ],
    "days": [
        ("周一 3/16", "run",   "跑步机跑 5.9km",    "45min", "HR135(Z3)", "配速7:40，心率已进Z3"),
        ("周二 3/17", "swim",  "游泳 1025m",         "53min", "HR131(Z3)", "1025m是本周最好游泳"),
        ("周三 3/18", "run",   "跑步机跑 4.38km",    "40min", "HR126(Z2)", "配速9:10，强度偏低"),
        ("周四 3/19", "swim",  "游泳 650m",          "41min", "HR133(Z3)", "距离偏短"),
        ("周五 3/20", "swim",  "游泳 575m（恢复？）", "66min", "HR101(Z1)", "HR极低，恢复性训练"),
        ("周六 3/21", "bike",  "骑行 30.7km",         "86min", "HR124(Z2)", "有氧骑，HR偏低"),
        ("周日 3/22", "brick", "Brick:骑20km+跑5km", "54+35min","骑HR129/跑HR157","首次Brick！跑步心率进Z4"),
    ]
}

# ════════════════════════════════════════════════════════════
# 未来两周详细计划
# ════════════════════════════════════════════════════════════

# 每个训练日格式：
# (日期, 星期, 运动类型, 训练名称, 总时长, [(阶段名, 时长min, 心率区间, 说明), ...], 关键指标, 注意事项)

WEEK1 = {
    "label": "W1（3/23-3/29）— 直接开干",
    "focus": "骑行心率提至Z3(130-148)，跑步阈值间歇首次引入，游泳750m基准测试",
    "days": [
        {
            "date": "3/23(周一)", "sport": "run", "name": "节奏跑 6km",
            "total": "40-45min",
            "phases": [
                ("热身慢跑",    8,  "Z2 111-130bpm", "放松配速8:00-9:00/km，心率不超130"),
                ("节奏跑",     24,  "Z3 130-148bpm", "配速目标6:20-6:40/km，保持心率130-145全程稳定"),
                ("放松慢跑",    8,  "Z2 111-130bpm", "降速至8:00+，心率回落至125以下"),
            ],
            "kpi": "节奏跑段平均HR 130-145，配速6:20-6:40稳定",
            "note": "跑步机或户外均可。若心率超150，降速保心率。",
        },
        {
            "date": "3/24(周二)", "sport": "swim", "name": "游泳间歇 4×200m + 200m计时",
            "total": "40-45min",
            "phases": [
                ("热身",        5,  "Z2 111-130bpm", "400m轻松自由泳，感受水感"),
                ("4×200m间歇", 16,  "Z3 130-148bpm", "每组200m目标<6:00，组间休息30s站立，专注换气节奏"),
                ("休息",        2,  "—",             "站立休息2min"),
                ("200m计时",    4,  "Z3-Z4 130-167", "全力出发，记录完整时间（目标<6:00，基准值）"),
                ("放松",        3,  "Z2",            "轻松游100-200m放松"),
            ],
            "kpi": "200m计时成绩记录（基准），每组心率130-148",
            "note": "25m泳池，每组记录时间。计时200m是本次核心数据。",
        },
        {
            "date": "3/25(周三)", "sport": "run", "name": "阈值间歇 3×1km @Z4",
            "total": "35-40min",
            "phases": [
                ("热身慢跑",   10,  "Z2 111-130bpm", "渐进加速，最后2min加至Z3"),
                ("1km阈值#1",   6,  "Z4 148-167bpm", "配速目标5:45-6:05/km，心率148-162"),
                ("慢跑恢复",  1.5,  "Z2",            "配速降至8:00+，心率回到130以下"),
                ("1km阈值#2",   6,  "Z4 148-167bpm", "同上，保持配速稳定"),
                ("慢跑恢复",  1.5,  "Z2",            ""),
                ("1km阈值#3",   6,  "Z4 148-167bpm", "最后一组可略微提速"),
                ("放松慢跑",    5,  "Z2",            "心率降回120以下再结束"),
            ],
            "kpi": "3组1km均保持Z4心率，配速5:45-6:05，组间HR回落至130以下",
            "note": "这是最重要的跑步课！间歇结束后充分拉伸股四头肌和小腿。",
        },
        {
            "date": "3/26(周四)", "sport": "rest", "name": "休息日",
            "total": "—",
            "phases": [("恢复", 0, "—", "充分休息，15min拉伸/泡沫轴。睡眠8h+")],
            "kpi": "—", "note": "不要强撑训练，明后天强度大。",
        },
        {
            "date": "3/27(周五)", "sport": "run", "name": "轻松跑 6km Z2",
            "total": "42-48min",
            "phases": [
                ("全程轻松跑", 42,  "Z2 111-130bpm", "全程控制心率111-130，配速7:00-7:30/km（慢也没关系）"),
            ],
            "kpi": "全程HR不超130，纯有氧基础恢复课",
            "note": "周四恢复后的激活课。心率超130就降速步行。",
        },
        {
            "date": "3/28(周六)", "sport": "brick", "name": "🧱 Brick：骑30km → 跑5km",
            "total": "骑80-90min + T1 + 跑35-38min",
            "phases": [
                ("热身骑行",   10,  "Z2 111-130bpm", "踏频≥85rpm，热身腿部"),
                ("强度骑行",   60,  "Z3-Z4 130-155", "目标HR 135-150！比上周(HR129)高10bpm！踏频85-90rpm"),
                ("冲刺段",     10,  "Z4 148-160",    "最后10min保持高HR，为换项做准备"),
                ("T1换项",      2,  "—",             "骑行停止→换跑鞋→出发，计时！目标<2min"),
                ("砖训跑步",   35,  "Z3 130-148bpm", "目标配速6:40-7:00/km，HR控制在148以下"),
            ],
            "kpi": "骑行均速HR>135（比上周高），跑步HR<150，T1<2min",
            "note": "骑行一定要提心率！上周HR124太低。跑步心率高是正常的，不要停。",
        },
        {
            "date": "3/29(周日)", "sport": "swim", "name": "🎯 750m基准计时",
            "total": "30-35min",
            "phases": [
                ("热身",        5,  "Z2",            "200m轻松游 + 2×50m渐进加速"),
                ("750m全力计时", 23, "Z3-Z4 130-167", "连续游750m，记录完整时间（全程不停）"),
                ("放松",        5,  "Z2",            "200m放松游"),
            ],
            "kpi": "🎯 750m完整计时成绩（基准值，目标<23min）",
            "note": "今天记录的成绩是未来9周进步参考点，重要！记录时间。",
        },
    ]
}

WEEK2 = {
    "label": "W2（3/30-4/5）— 强度巩固",
    "focus": "跑步阈值增至4组，游泳超距800m，骑行Brick强度提至HR140-155",
    "days": [
        {
            "date": "3/30(周一)", "sport": "run", "name": "长跑 8km Z2",
            "total": "55-60min",
            "phases": [
                ("全程有氧长跑", 56, "Z2 111-130bpm", "稳定配速7:00-7:30/km，心率控制在125以下，跑最长单次距离"),
            ],
            "kpi": "8km完成，全程HR<130，有氧基础",
            "note": "本周最长跑，配速不重要，心率不超130。跑前喝水。",
        },
        {
            "date": "3/31(周二)", "sport": "swim", "name": "2×400m Z3 + 100m放松",
            "total": "40-45min",
            "phases": [
                ("热身",       5,  "Z2",            "200m轻松 + 4×25m渐进"),
                ("400m#1",    12,  "Z3 130-148",    "连续400m，配速目标3:00/100m，记录时间"),
                ("组间休息",   2,  "—",             "站立或慢走2min，心率降至120以下"),
                ("400m#2",    12,  "Z3 130-148",    "配速与第1组相近或更快"),
                ("100m放松",   3,  "Z2",            "放松游100m"),
            ],
            "kpi": "两组400m时间记录，均速配速<3:00/100m",
            "note": "400m是本周游泳强度核心，记录每组时间对比。",
        },
        {
            "date": "4/1(周三)", "sport": "run", "name": "阈值间歇 4×1km @Z4",
            "total": "45-50min",
            "phases": [
                ("热身慢跑",   10, "Z2 111-130bpm", "渐进加速"),
                ("1km阈值×4",  24, "Z4 148-167bpm", "每组6min，配速5:45-6:05，组间慢跑90s（HR降至130以下）"),
                ("慢跑恢复×4", 6,  "Z2",            "4次×1.5min组间恢复"),
                ("放松慢跑",   5,  "Z2",            ""),
            ],
            "kpi": "4组均保持Z4心率，较W1多1组，配速稳定",
            "note": "比上周多1组，若完成后腿很沉可缩减第4组至800m。",
        },
        {
            "date": "4/2(周四)", "sport": "rest", "name": "休息日",
            "total": "—",
            "phases": [("恢复", 0, "—", "拉伸+泡沫轴20min，充分睡眠")],
            "kpi": "—", "note": "—",
        },
        {
            "date": "4/3(周五)", "sport": "run", "name": "节奏跑 6km",
            "total": "40-45min",
            "phases": [
                ("热身慢跑",    8, "Z2 111-130bpm", ""),
                ("节奏跑",     24, "Z3 130-148bpm", "配速6:20-6:40，比W1节奏跑感觉更轻松说明体能在提升"),
                ("放松慢跑",    8, "Z2",            ""),
            ],
            "kpi": "节奏跑段HR130-145，配速6:20-6:40",
            "note": "对比W1的感觉，应该更轻松。",
        },
        {
            "date": "4/4(周六)", "sport": "brick", "name": "🧱 Brick：骑35km → 跑5km",
            "total": "骑95-105min + T1 + 跑35min",
            "phases": [
                ("热身骑行",   10, "Z2",            "踏频≥85rpm"),
                ("强度骑行",   70, "Z3-Z4 135-155", "目标HR 140-152，比上周再高！踏频85-90rpm"),
                ("冲刺段",     15, "Z4",            "最后15min维持HR 148+"),
                ("T1换项",      2, "—",             "计时，目标<1:50（比上周快）"),
                ("砖训跑步",   35, "Z3 130-148",    "目标配速6:30-6:50/km，控制HR在145以下"),
            ],
            "kpi": "骑行均HR>140，跑步配速<6:50，T1<1:50",
            "note": "骑行距离+5km，强度提升。跑步目标比W1快10-20s/km。",
        },
        {
            "date": "4/5(周日)", "sport": "swim", "name": "超距 800m 连续游",
            "total": "35-40min",
            "phases": [
                ("热身",        5, "Z2",            "200m轻松"),
                ("800m超距",   24, "Z2-Z3 111-148", "连续游800m不停，比赛距离750m的基础，配速3:00/100m"),
                ("放松",        5, "Z2",            "200m放松"),
            ],
            "kpi": "800m完成时间（目标<25min），全程不停",
            "note": "比比赛距离多50m，建立心理余量。记录时间对比下周进步。",
        },
    ]
}

SPORT_BG = {"run": C_RUN, "swim": C_SWIM, "bike": C_BIKE, "brick": C_BRICK, "rest": C_REST}
SPORT_EMOJI = {"run": "🏃", "swim": "🏊", "bike": "🚴", "brick": "🧱", "rest": "😴"}

def generate():
    wb = openpyxl.Workbook()

    # ── Sheet 1：本周复盘 ──────────────────────────────────
    ws0 = wb.active
    ws0.title = "本周复盘 3.16-3.22"

    ws0.merge_cells("A1:F1")
    c = ws0.cell(row=1, column=1, value=f"📊 {ACTUAL_WEEK['label']}")
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.fill = fill(C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = tb()
    ws0.row_dimensions[1].height = 26

    # 总结
    ws0.merge_cells("A2:F2")
    wc(ws0, 2, 1, ACTUAL_WEEK["summary"], bg="FFF2CC", sz=10)
    ws0.row_dimensions[2].height = 40

    # 亮点/问题
    row = 3
    for h in ACTUAL_WEEK["highlights"]:
        ws0.merge_cells(f"A{row}:F{row}")
        bg = C_WARN if h.startswith("⚠️") else "E2EFDA"
        wc(ws0, row, 1, h, bg=bg, sz=9)
        ws0.row_dimensions[row].height = 20
        row += 1

    row += 1
    headers = ["日期", "项目", "训练内容", "时长", "心率", "备注"]
    widths   = [14, 8, 22, 12, 18, 28]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        wc(ws0, row, i, h, bold=True, fc="FFFFFF", bg=C_HEADER, ha="center")
        ws0.column_dimensions[get_column_letter(i)].width = w
    ws0.row_dimensions[row].height = 18
    row += 1

    for day in ACTUAL_WEEK["days"]:
        date_s, sport, content, dur, hr, note = day
        bg = SPORT_BG.get(sport, "FFFFFF")
        emoji = SPORT_EMOJI.get(sport, "")
        wc(ws0, row, 1, date_s, bg=bg, ha="center")
        wc(ws0, row, 2, emoji + " " + sport, bg=bg, ha="center")
        wc(ws0, row, 3, content, bg=bg)
        wc(ws0, row, 4, dur, bg=bg, ha="center")
        wc(ws0, row, 5, hr, bg=bg, ha="center")
        wc(ws0, row, 6, note, bg=bg)
        ws0.row_dimensions[row].height = 30
        row += 1

    # ── Sheet 2-3：未来两周详细计划 ───────────────────────
    for week_idx, week_data in enumerate([WEEK1, WEEK2], 1):
        ws = wb.create_sheet(title=f"W{week_idx}详细计划")
        ws.freeze_panes = "B3"

        ws.merge_cells("A1:H1")
        c = ws.cell(row=1, column=1,
            value=f"🏋️ {week_data['label']}  |  📌 {week_data['focus']}")
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = fill(C_HEADER)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = tb()
        ws.row_dimensions[1].height = 36

        headers = ["日期/运动", "训练阶段", "时长(min)", "心率区间", "配速/强度要求", "关键指标", "注意事项"]
        widths   = [22, 18, 10, 18, 30, 28, 30]
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            wc(ws, 2, i, h, bold=True, fc="FFFFFF", bg=C_HEADER, ha="center")
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[2].height = 18

        row = 3
        for day in week_data["days"]:
            sport = day["sport"]
            bg = SPORT_BG.get(sport, "FFFFFF")
            emoji = SPORT_EMOJI.get(sport, "")
            phases = day["phases"]
            n = len(phases)

            # 日期列合并
            if n > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row+n-1, end_column=1)
            mc = ws.cell(row=row, column=1,
                value=f"{emoji} {day['date']}\n{day['name']}\n总时长:{day['total']}")
            mc.font = Font(bold=True, size=9, color="1F3864")
            mc.fill = fill(bg)
            mc.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
            mc.border = tb()

            # KPI和注意事项合并
            if n > 1:
                ws.merge_cells(start_row=row, start_column=6, end_row=row+n-1, end_column=6)
                ws.merge_cells(start_row=row, start_column=7, end_row=row+n-1, end_column=7)

            wc(ws, row, 6, day["kpi"], bg="FFFACD", sz=9)
            wc(ws, row, 7, day["note"], bg="F0F8FF", sz=9)

            for pi, (pname, pdur, phr, pdesc) in enumerate(phases):
                r = row + pi
                alt = "F9F9F9" if pi % 2 == 1 else bg
                if pname == "恢复" and sport == "rest":
                    wc(ws, r, 2, "完全休息", bg=C_REST, ha="center")
                    wc(ws, r, 3, "全天", bg=C_REST, ha="center")
                    wc(ws, r, 4, "—", bg=C_REST, ha="center")
                    wc(ws, r, 5, pdesc, bg=C_REST)
                else:
                    wc(ws, r, 2, pname, bg=alt, bold=(pi==0))
                    wc(ws, r, 3, str(pdur) if pdur else "—", bg=alt, ha="center")
                    wc(ws, r, 4, phr, bg=alt, ha="center")
                    wc(ws, r, 5, pdesc, bg=alt)
                ws.row_dimensions[r].height = 36

            row += n
            # 分隔行
            ws.row_dimensions[row].height = 6
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill("DDDDDD")
            row += 1

    wb.save(OUTPUT)
    print(f"✅ Excel 已生成：{OUTPUT}")

if __name__ == "__main__":
    generate()
